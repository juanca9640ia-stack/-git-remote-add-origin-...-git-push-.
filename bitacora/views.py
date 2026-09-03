import uuid
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from inventario.models import Categoria, Producto
from ventas.models import Cotizacion, CuentaCobro, LineaCotizacion

from .forms import ItemBitacoraForm, SedeForm
from .models import ItemBitacora, Sede


@login_required
def sede_lista(request):
    sedes = Sede.objects.select_related("cliente").filter(empresa=request.empresa)
    cliente_id = request.GET.get("cliente", "")
    if cliente_id:
        sedes = sedes.filter(cliente_id=cliente_id)
    query = request.GET.get("q", "")
    if query:
        sedes = sedes.filter(nombre__icontains=query)
    return render(request, "bitacora/sede_lista.html", {
        "sedes": sedes, "query": query, "cliente_id": cliente_id,
        "nueva_sede_form": SedeForm(empresa=request.empresa),
    })


@login_required
def sede_form(request, pk=None):
    sede = get_object_or_404(Sede, pk=pk, empresa=request.empresa) if pk else None
    if request.method == "POST":
        form = SedeForm(request.POST, instance=sede, empresa=request.empresa)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.empresa = request.empresa
            obj.save()
            messages.success(request, f"Sede '{obj.nombre}' guardada correctamente.")
            # El alta rápida desde la lista de sedes, desde la ficha del cliente, o
            # desde la hoja de otra sede del mismo cliente, se queda en esa misma
            # hoja para poder seguir agregando sedes seguidas en vez de saltar al
            # detalle de la sede recién creada.
            origen = request.POST.get("origen")
            if not sede and origen == "lista":
                return redirect("bitacora:sede_lista")
            if not sede and origen == "cliente":
                return redirect("ventas:cliente_detalle", pk=obj.cliente_id)
            if not sede and origen == "sede":
                volver_pk = request.POST.get("volver") or obj.pk
                return redirect("bitacora:sede_detalle", pk=volver_pk)
            return redirect("bitacora:sede_detalle", pk=obj.pk)
    else:
        initial = {}
        if request.GET.get("cliente"):
            initial["cliente"] = request.GET["cliente"]
        form = SedeForm(instance=sede, empresa=request.empresa, initial=initial)
    return render(request, "bitacora/sede_form.html", {"form": form, "sede": sede})


def _rango_por_defecto(request):
    """Mes actual por defecto; se puede ampliar con ?desde=&hasta= o ver todo con ?todo=1."""
    hoy = timezone.localdate()
    if request.GET.get("todo") == "1":
        return None, None
    desde = request.GET.get("desde") or hoy.replace(day=1).isoformat()
    hasta = request.GET.get("hasta") or hoy.isoformat()
    return desde, hasta


@login_required
def sede_detalle(request, pk):
    """Hoja de trabajo diario del cliente completo: todas sus sedes, una tras
    otra en la misma página (cada una con su propia alta de ítems y su propia
    tabla), tal como venía la hoja de cálculo original. `pk` solo identifica
    con qué sede se entró, para saber a cuál volver tras un alta rápida."""
    sede = get_object_or_404(Sede.objects.select_related("cliente"), pk=pk, empresa=request.empresa)
    cliente = sede.cliente
    desde, hasta = _rango_por_defecto(request)

    sedes_cliente = Sede.objects.filter(cliente=cliente, empresa=request.empresa).order_by("nombre")

    bloques = []
    for s in sedes_cliente:
        items = s.items.all()
        if desde:
            items = items.filter(fecha__gte=desde)
        if hasta:
            items = items.filter(fecha__lte=hasta)
        items = list(items)
        subtotal = sum((i.subtotal for i in items), Decimal("0"))
        iva_total = sum((i.iva for i in items), Decimal("0"))
        pendientes = [i for i in items if not i.facturado]
        bloques.append({
            "sede": s,
            "items": items,
            "subtotal": subtotal,
            "iva_total": iva_total,
            "total_con_iva": subtotal + iva_total,
            "pendientes_count": len(pendientes),
            "subtotal_pendiente": sum((i.subtotal for i in pendientes), Decimal("0")),
            # auto_id propio por sede: son varios <form> de alta de ítem en la
            # misma página y los id= de sus campos no pueden repetirse.
            "item_form": ItemBitacoraForm(
                initial={"fecha": timezone.localdate()}, auto_id=f"id_item_{s.pk}_%s",
            ),
        })

    return render(request, "bitacora/sede_detalle.html", {
        "cliente": cliente, "sede": sede, "sedes": sedes_cliente, "bloques": bloques,
        "desde": desde, "hasta": hasta, "ver_todo": desde is None,
    })


@login_required
@require_POST
def item_crear(request, pk):
    sede = get_object_or_404(Sede, pk=pk, empresa=request.empresa)
    form = ItemBitacoraForm(request.POST)
    if form.is_valid():
        item = form.save(commit=False)
        item.empresa = request.empresa
        item.sede = sede
        item.creado_por = request.user
        item.save()
        messages.success(request, "Ítem agregado a la bitácora.")
    else:
        messages.error(request, "Revisa los datos del ítem: " + "; ".join(
            f"{campo}: {', '.join(errores)}" for campo, errores in form.errors.items()
        ))
    return redirect("bitacora:sede_detalle", pk=sede.pk)


@login_required
@require_POST
def item_eliminar(request, pk, item_pk):
    sede = get_object_or_404(Sede, pk=pk, empresa=request.empresa)
    item = get_object_or_404(ItemBitacora, pk=item_pk, sede=sede)
    if item.facturado:
        messages.error(request, "Este ítem ya fue incluido en un documento y no se puede eliminar.")
    else:
        item.delete()
        messages.success(request, "Ítem eliminado.")
    return redirect("bitacora:sede_detalle", pk=sede.pk)


@login_required
def sede_exportar_excel(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    sede = get_object_or_404(Sede.objects.select_related("cliente"), pk=pk, empresa=request.empresa)
    desde, hasta = _rango_por_defecto(request)
    items = sede.items.all()
    if desde:
        items = items.filter(fecha__gte=desde)
    if hasta:
        items = items.filter(fecha__lte=hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = sede.nombre[:31] or "Bitácora"

    ws.merge_cells("A1:I1")
    ws["A1"] = sede.nombre.upper()
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([f"{sede.cliente}"])

    encabezados = [
        "Item", "Mantenimiento general", "Unidad", "Cantidad", "Vr.uni",
        "Sub total", "IVA", "Vr.total", "Estado", "Notas",
    ]
    ws.append(encabezados)
    for cell in ws[3]:
        cell.font = Font(bold=True)

    subtotal = Decimal("0")
    iva_total = Decimal("0")
    for indice, item in enumerate(items, start=1):
        ws.append([
            indice, item.descripcion, item.unidad, float(item.cantidad), float(item.valor_unitario),
            float(item.subtotal), float(item.iva), float(item.valor_total),
            "Facturado" if item.facturado else "Pendiente", item.notas,
        ])
        subtotal += item.subtotal
        iva_total += item.iva

    fila_total = ws.max_row + 1
    ws.cell(row=fila_total, column=5, value="Subtotal").font = Font(bold=True)
    ws.cell(row=fila_total, column=6, value=float(subtotal)).font = Font(bold=True)
    ws.cell(row=fila_total, column=7, value=float(iva_total)).font = Font(bold=True)
    ws.cell(row=fila_total, column=8, value=float(subtotal + iva_total)).font = Font(bold=True)

    anchos = [6, 40, 8, 10, 14, 14, 12, 14, 12, 30]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nombre_archivo = f"bitacora_{sede.nombre}_{timezone.localdate().isoformat()}.xlsx".replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@login_required
def sede_exportar_pdf(request, pk):
    sede = get_object_or_404(Sede.objects.select_related("cliente"), pk=pk, empresa=request.empresa)
    desde, hasta = _rango_por_defecto(request)
    items = sede.items.all()
    if desde:
        items = items.filter(fecha__gte=desde)
    if hasta:
        items = items.filter(fecha__lte=hasta)
    items = list(items)
    subtotal = sum((i.subtotal for i in items), Decimal("0"))
    iva_total = sum((i.iva for i in items), Decimal("0"))
    return render(request, "bitacora/sede_pdf.html", {
        "sede": sede, "items": items, "subtotal": subtotal, "iva_total": iva_total,
        "total": subtotal + iva_total, "empresa": request.empresa, "desde": desde, "hasta": hasta,
    })


@login_required
@require_POST
def sede_generar_cuenta_cobro(request, pk):
    sede = get_object_or_404(Sede.objects.select_related("cliente"), pk=pk, empresa=request.empresa)
    pendientes = list(sede.items.filter(cotizacion__isnull=True, cuenta_cobro__isnull=True))
    if not pendientes:
        messages.error(request, "No hay ítems pendientes por facturar en esta sede.")
        return redirect("bitacora:sede_detalle", pk=sede.pk)

    lineas = "\n".join(
        f"- {i.fecha.strftime('%d/%m/%Y')}: {i.descripcion} "
        f"({i.cantidad} {i.unidad} x ${i.valor_unitario:,.0f} = ${i.subtotal:,.0f})"
        for i in pendientes
    )
    total = sum((i.subtotal for i in pendientes), Decimal("0"))

    with transaction.atomic():
        cuenta = CuentaCobro.objects.create(
            empresa=request.empresa, cliente=sede.cliente,
            concepto=f"Bitácora de trabajo diario — sede {sede.nombre}:\n{lineas}",
            valor=total, creado_por=request.user,
        )
        ItemBitacora.objects.filter(pk__in=[i.pk for i in pendientes]).update(cuenta_cobro=cuenta)

    messages.success(request, f"Cuenta de cobro '{cuenta.numero}' generada con {len(pendientes)} ítem(s).")
    return redirect("ventas:cuenta_cobro_detalle", pk=cuenta.pk)


@login_required
@require_POST
@transaction.atomic
def sede_generar_cotizacion(request, pk):
    sede = get_object_or_404(Sede.objects.select_related("cliente"), pk=pk, empresa=request.empresa)
    pendientes = list(sede.items.filter(cotizacion__isnull=True, cuenta_cobro__isnull=True))
    if not pendientes:
        messages.error(request, "No hay ítems pendientes por cotizar en esta sede.")
        return redirect("bitacora:sede_detalle", pk=sede.pk)

    categoria_servicios, _ = Categoria.objects.get_or_create(
        empresa=request.empresa, nombre="Servicios de bitácora",
    )

    cotizacion = Cotizacion.objects.create(
        empresa=request.empresa, cliente=sede.cliente, vendedor=request.user,
        notas=f"Generada desde la bitácora de trabajo diario de la sede {sede.nombre}.",
    )

    for item in pendientes:
        producto = Producto.objects.filter(
            empresa=request.empresa, nombre__iexact=item.descripcion, tipo=Producto.SERVICIO,
        ).first()
        if not producto:
            producto = Producto.objects.create(
                empresa=request.empresa, sku=f"BIT-{uuid.uuid4().hex[:8].upper()}",
                nombre=item.descripcion[:150], tipo=Producto.SERVICIO,
                categoria=categoria_servicios, precio_venta=item.valor_unitario,
            )
        LineaCotizacion.objects.create(
            empresa=request.empresa, cotizacion=cotizacion, producto=producto,
            cantidad=item.cantidad, precio_unitario=item.valor_unitario,
        )

    ItemBitacora.objects.filter(pk__in=[i.pk for i in pendientes]).update(cotizacion=cotizacion)

    messages.success(request, f"Cotización {cotizacion.numero} generada con {len(pendientes)} ítem(s).")
    return redirect("ventas:cotizacion_detalle", pk=cotizacion.pk)
